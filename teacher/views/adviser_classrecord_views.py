# teacher/views/adviser_classrecord_views.py

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.db import transaction
from django.core.exceptions import ValidationError
from django.utils import timezone
from django.db.models import Q, Prefetch
from datetime import datetime
import json

# Import models from teacher app
from teacher.models import (
    ClassRecord, StudentGrade, GradeSummary,
    MasterlistStudent, AdviserMasterlist, EarlyWarningAlert
)

# Import models from admin_functionalities app
from admin_functionalities.models import (
    Teacher, Section, Subject, SectionSubjectAssignment,
    SchoolYear
)

# Import models from enrollmentprocess app
from enrollmentprocess.models import Student


# ============================================================================
# MAIN VIEW - Class Record Page
# ============================================================================

@login_required
def adviser_classrecord(request):
    """
    Main view for the E-Class Record page.
    Displays the interface for managing student grades across quarters.
    """
    try:
        # Get the logged-in teacher
        teacher = Teacher.objects.get(user=request.user)
        
        # Get current school year
        current_sy = SchoolYear.get_current()
        if not current_sy:
            return render(request, 'teacher/adviser/Class_Record.html', {
                'error': 'No active school year found. Please contact administrator.',
                'teacher': teacher,
                'all_sections': [],
                'all_subjects': [],
                'school_year_choices': [],
                'sections_with_subjects': '{}',
                'quarter_choices': ClassRecord.QUARTER_CHOICES,
            })
        
        # CRITICAL: Get sections where teacher is adviser
        adviser_sections = Section.objects.filter(
            adviser=teacher,
            is_active=True
        )
        
        # CRITICAL: Get sections where teacher has subject assignments
        taught_section_ids = SectionSubjectAssignment.objects.filter(
            teacher=teacher
        ).values_list('section_id', flat=True).distinct()
        
        taught_sections = Section.objects.filter(
            id__in=taught_section_ids,
            is_active=True
        )
        
        # Combine sections
        all_sections = (adviser_sections | taught_sections).distinct().order_by('name')
        
        # Debug output
        print("\n" + "="*80)
        print(f"DEBUG - Teacher: {teacher.full_name} (ID: {teacher.id})")
        print(f"DEBUG - Adviser sections: {adviser_sections.count()}")
        for sec in adviser_sections:
            print(f"  - {sec.name} (ID: {sec.id})")
        print(f"DEBUG - Taught sections: {taught_sections.count()}")
        for sec in taught_sections:
            print(f"  - {sec.name} (ID: {sec.id})")
        print(f"DEBUG - Total unique sections: {all_sections.count()}")
        print("="*80 + "\n")
        
        # Get subject assignments
        subject_assignments = SectionSubjectAssignment.objects.filter(
            teacher=teacher
        ).select_related('section')
        
        print(f"DEBUG - Subject assignments: {subject_assignments.count()}")
        for assignment in subject_assignments:
            print(f"  - {assignment.section.name}: {assignment.subject}")
        
        # Organize subjects by section
        sections_with_subjects = {}
        for assignment in subject_assignments:
            section_id = str(assignment.section.id)
            if section_id not in sections_with_subjects:
                sections_with_subjects[section_id] = {
                    'section_id': assignment.section.id,
                    'section_name': assignment.section.name,
                    'program': assignment.section.program,
                    'subjects': []
                }
            
            # Get subject display name from choices
            subject_choices_dict = dict(SectionSubjectAssignment._meta.get_field('subject').choices)
            subject_display = subject_choices_dict.get(assignment.subject, assignment.subject)
            
            sections_with_subjects[section_id]['subjects'].append({
                'code': assignment.subject,
                'name': subject_display,
                'assignment_id': assignment.id
            })
        
        print(f"\nDEBUG - Sections with subjects structure:")
        import pprint
        pprint.pprint(sections_with_subjects)
        print("\n")
        
        # Get all subjects for dropdown
        all_subjects = Subject.objects.filter(is_active=True).order_by('display_order', 'subject_name')
        
        # Get school years
        school_year_choices = SchoolYear.objects.filter(is_active=True).order_by('-start_date')
        
        # Convert to JSON
        sections_with_subjects_json = json.dumps(sections_with_subjects)
        
        print(f"DEBUG - JSON output length: {len(sections_with_subjects_json)}")
        print(f"DEBUG - JSON preview: {sections_with_subjects_json[:200]}...\n")
        
        context = {
            'teacher': teacher,
            'current_school_year': current_sy,
            'school_year_choices': school_year_choices,
            'all_sections': all_sections,
            'adviser_sections': adviser_sections,
            'taught_sections': taught_sections,
            'sections_with_subjects': sections_with_subjects_json,
            'all_subjects': all_subjects,
            'quarter_choices': ClassRecord.QUARTER_CHOICES,
        }
        
        return render(request, 'teacher/adviser/Class_Record.html', context)
        
    except Teacher.DoesNotExist:
        print("ERROR: Teacher profile not found!")
        return render(request, 'teacher/adviser/Class_Record.html', {
            'error': 'Teacher profile not found. Please contact administrator.',
            'all_sections': [],
            'all_subjects': [],
            'school_year_choices': [],
            'sections_with_subjects': '{}',
            'quarter_choices': [],
        })
    except Exception as e:
        print(f"ERROR in adviser_classrecord: {str(e)}")
        import traceback
        traceback.print_exc()
        return render(request, 'teacher/adviser/Class_Record.html', {
            'error': f'An error occurred: {str(e)}',
            'all_sections': [],
            'all_subjects': [],
            'school_year_choices': [],
            'sections_with_subjects': '{}',
            'quarter_choices': [],
        })

# ============================================================================
# API ENDPOINTS - Class Record CRUD Operations
# ============================================================================

@login_required
@require_http_methods(["GET"])
def get_class_record(request):
    """
    Get or create a class record for a specific section, subject, quarter, and school year.
    Returns class record configuration and student grades.
    """
    try:
        teacher = Teacher.objects.get(user=request.user)
        
        # Get parameters - ACCEPT BOTH subject_code AND subject_id for compatibility
        section_id = request.GET.get('section_id')
        subject_code = request.GET.get('subject_code') or request.GET.get('subject_id')  # <-- FIX HERE
        quarter = request.GET.get('quarter')
        school_year = request.GET.get('school_year')
        
        print(f"\n{'='*80}")
        print(f"GET CLASS RECORD REQUEST")
        print(f"Section ID: {section_id}")
        print(f"Subject Code: {subject_code}")
        print(f"Quarter: {quarter}")
        print(f"School Year: {school_year}")
        print(f"{'='*80}\n")
        
        # Validate required parameters
        if not all([section_id, subject_code, quarter, school_year]):
            missing = []
            if not section_id: missing.append('section_id')
            if not subject_code: missing.append('subject_code/subject_id')
            if not quarter: missing.append('quarter')
            if not school_year: missing.append('school_year')
            
            return JsonResponse({
                'success': False,
                'error': f'Missing required parameters: {", ".join(missing)}'
            }, status=400)
        
        # Get related objects
        section = get_object_or_404(Section, id=section_id, is_active=True)
        
        # Find the subject assignment to get the actual Subject
        assignment = SectionSubjectAssignment.objects.filter(
            section=section,
            subject=subject_code,
            teacher=teacher
        ).first()
        
        if not assignment:
            return JsonResponse({
                'success': False,
                'error': f'No assignment found for {subject_code} in {section.name}'
            }, status=404)
        
        # Get or create the Subject model
        subject = Subject.objects.filter(subject_code=subject_code).first()
        
        if not subject:
            # Create a Subject if it doesn't exist
            subject_choices_dict = dict(SectionSubjectAssignment._meta.get_field('subject').choices)
            subject_name = subject_choices_dict.get(subject_code, subject_code)
            
            subject = Subject.objects.create(
                subject_code=subject_code,
                subject_name=subject_name,
                subject_type='CORE',
                is_active=True
            )
            print(f"Created new Subject: {subject}")
        
        # Check if teacher has access to this section/subject
        has_access = (
            section.adviser == teacher or
            assignment is not None
        )
        
        if not has_access:
            return JsonResponse({
                'success': False,
                'error': 'You do not have access to this class record'
            }, status=403)
        
        # Get or create class record
        class_record, created = ClassRecord.objects.get_or_create(
            teacher=teacher,
            subject=subject,
            section=section,
            quarter=quarter,
            school_year=school_year,
            defaults={
                'written_works_weight': 30,
                'performance_tasks_weight': 50,
                'quarterly_assessment_weight': 20,
            }
        )
        
        print(f"Class Record: {class_record} ({'created' if created else 'existing'})")
        
        # ========================================================================
        # FETCH STUDENTS FROM ADVISER MASTERLIST (SAME AS ATTENDANCE VIEW)
        # ========================================================================
        
        from teacher.models import AdviserMasterlist, MasterlistStudent
        
        masterlist = AdviserMasterlist.objects.filter(
            section=section,
            school_year=school_year,
            is_active=True
        ).first()
        
        students_data = []
        
        if masterlist:
            print(f"Found Masterlist: {masterlist}")
            print(f"Total students in masterlist: {masterlist.total_students}")
            
            # Get active students from masterlist
            masterlist_students = MasterlistStudent.objects.filter(
                masterlist=masterlist,
                is_active=True
            ).select_related('student').order_by('student__last_name', 'student__first_name')
            
            print(f"Active students found: {masterlist_students.count()}")
            
            # Get or create student grades
            for idx, ms in enumerate(masterlist_students, start=1):
                student = ms.student
                
                print(f"  {idx}. {student.last_name}, {student.first_name} (ID: {student.id})")
                
                # Get or create student grade
                student_grade, grade_created = StudentGrade.objects.get_or_create(
                    class_record=class_record,
                    student=student
                )
                
                students_data.append({
                    'id': student.id,
                    'number': idx,
                    'name': f"{student.last_name}, {student.first_name} {student.middle_name or ''}".strip(),
                    'gender': student.gender,
                    'grade_id': student_grade.id,
                    'scores': {
                        'ww': [
                            student_grade.ww_score_1, student_grade.ww_score_2,
                            student_grade.ww_score_3, student_grade.ww_score_4,
                            student_grade.ww_score_5, student_grade.ww_score_6,
                            student_grade.ww_score_7, student_grade.ww_score_8,
                            student_grade.ww_score_9, student_grade.ww_score_10
                        ],
                        'pt': [
                            student_grade.pt_score_1, student_grade.pt_score_2,
                            student_grade.pt_score_3, student_grade.pt_score_4,
                            student_grade.pt_score_5, student_grade.pt_score_6,
                            student_grade.pt_score_7, student_grade.pt_score_8,
                            student_grade.pt_score_9, student_grade.pt_score_10
                        ],
                        'qa': [student_grade.qa_score_1]
                    },
                    'computed': {
                        'ww_total': student_grade.ww_total,
                        'ww_percentage': student_grade.ww_percentage,
                        'ww_weighted_score': student_grade.ww_weighted_score,
                        'pt_total': student_grade.pt_total,
                        'pt_percentage': student_grade.pt_percentage,
                        'pt_weighted_score': student_grade.pt_weighted_score,
                        'qa_percentage': student_grade.qa_percentage,
                        'qa_weighted_score': student_grade.qa_weighted_score,
                        'initial_grade': student_grade.initial_grade,
                        'quarterly_grade': student_grade.quarterly_grade
                    }
                })
        else:
            # Fallback: Try to get from SectionPlacement if no masterlist exists
            print(f"No masterlist found for {section.name} - {school_year}")
            print(f"Trying SectionPlacement as fallback...")
            
            from enrollmentprocess.models import SectionPlacement
            
            placements = SectionPlacement.objects.filter(
                section=section,
                status='approved'
            ).select_related('student').order_by('student__last_name', 'student__first_name')
            
            print(f"Students from SectionPlacement: {placements.count()}")
            
            for idx, placement in enumerate(placements, start=1):
                student = placement.student
                
                print(f"  {idx}. {student.last_name}, {student.first_name} (ID: {student.id})")
                
                # Get or create student grade
                student_grade, _ = StudentGrade.objects.get_or_create(
                    class_record=class_record,
                    student=student
                )
                
                students_data.append({
                    'id': student.id,
                    'number': idx,
                    'name': f"{student.last_name}, {student.first_name} {student.middle_name or ''}".strip(),
                    'gender': student.gender,
                    'grade_id': student_grade.id,
                    'scores': {
                        'ww': [
                            student_grade.ww_score_1, student_grade.ww_score_2,
                            student_grade.ww_score_3, student_grade.ww_score_4,
                            student_grade.ww_score_5, student_grade.ww_score_6,
                            student_grade.ww_score_7, student_grade.ww_score_8,
                            student_grade.ww_score_9, student_grade.ww_score_10
                        ],
                        'pt': [
                            student_grade.pt_score_1, student_grade.pt_score_2,
                            student_grade.pt_score_3, student_grade.pt_score_4,
                            student_grade.pt_score_5, student_grade.pt_score_6,
                            student_grade.pt_score_7, student_grade.pt_score_8,
                            student_grade.pt_score_9, student_grade.pt_score_10
                        ],
                        'qa': [student_grade.qa_score_1]
                    },
                    'computed': {
                        'ww_total': student_grade.ww_total,
                        'ww_percentage': student_grade.ww_percentage,
                        'ww_weighted_score': student_grade.ww_weighted_score,
                        'pt_total': student_grade.pt_total,
                        'pt_percentage': student_grade.pt_percentage,
                        'pt_weighted_score': student_grade.pt_weighted_score,
                        'qa_percentage': student_grade.qa_percentage,
                        'qa_weighted_score': student_grade.qa_weighted_score,
                        'initial_grade': student_grade.initial_grade,
                        'quarterly_grade': student_grade.quarterly_grade
                    }
                })
        
        print(f"\nTotal students loaded: {len(students_data)}\n")
        
        if len(students_data) == 0:
            print("⚠ WARNING: No students found!")
            print(f"  - Masterlist exists: {masterlist is not None}")
            if masterlist:
                print(f"  - Masterlist total_students: {masterlist.total_students}")
                print(f"  - Active MasterlistStudent count: {MasterlistStudent.objects.filter(masterlist=masterlist, is_active=True).count()}")
        
        # Prepare response data
        response_data = {
            'success': True,
            'created': created,
            'class_record': {
                'id': class_record.id,
                'section': section.name,
                'subject': subject.subject_name,
                'quarter': quarter,
                'school_year': school_year,
                'weights': {
                    'ww': class_record.written_works_weight,
                    'pt': class_record.performance_tasks_weight,
                    'qa': class_record.quarterly_assessment_weight
                },
                'hps': {
                    'ww': [
                        class_record.ww_hps_1, class_record.ww_hps_2,
                        class_record.ww_hps_3, class_record.ww_hps_4,
                        class_record.ww_hps_5, class_record.ww_hps_6,
                        class_record.ww_hps_7, class_record.ww_hps_8,
                        class_record.ww_hps_9, class_record.ww_hps_10
                    ],
                    'pt': [
                        class_record.pt_hps_1, class_record.pt_hps_2,
                        class_record.pt_hps_3, class_record.pt_hps_4,
                        class_record.pt_hps_5, class_record.pt_hps_6,
                        class_record.pt_hps_7, class_record.pt_hps_8,
                        class_record.pt_hps_9, class_record.pt_hps_10
                    ],
                    'qa': [class_record.qa_hps_1]
                },
                'hps_totals': {
                    'ww': class_record.get_ww_hps_total(),
                    'pt': class_record.get_pt_hps_total(),
                    'qa': class_record.get_qa_hps_total()
                }
            },
            'students': students_data
        }
        
        print(f"✅ Returning response with {len(students_data)} students\n")
        
        return JsonResponse(response_data)
        
    except Teacher.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Teacher profile not found'
        }, status=404)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@require_http_methods(["POST"])
def save_class_record(request):
    """
    Save class record configuration and all student grades.
    Handles bulk update of scores, HPS, and weights.
    """
    try:
        teacher = Teacher.objects.get(user=request.user)
        data = json.loads(request.body)
        
        class_record_id = data.get('class_record_id')
        if not class_record_id:
            return JsonResponse({
                'success': False,
                'error': 'Class record ID is required'
            }, status=400)
        
        # Get class record and verify ownership
        class_record = get_object_or_404(ClassRecord, id=class_record_id, teacher=teacher)
        
        with transaction.atomic():
            # Update grading criteria weights
            if 'weights' in data:
                weights = data['weights']
                class_record.written_works_weight = weights.get('ww', 30)
                class_record.performance_tasks_weight = weights.get('pt', 50)
                class_record.quarterly_assessment_weight = weights.get('qa', 20)
            
            # Update HPS values
            if 'hps' in data:
                hps = data['hps']
                if 'ww' in hps:
                    for i, value in enumerate(hps['ww'], start=1):
                        setattr(class_record, f'ww_hps_{i}', value)
                if 'pt' in hps:
                    for i, value in enumerate(hps['pt'], start=1):
                        setattr(class_record, f'pt_hps_{i}', value)
                if 'qa' in hps:
                    class_record.qa_hps_1 = hps['qa'][0]
            
            class_record.save()
            
            # Update student grades
            if 'students' in data:
                for student_data in data['students']:
                    grade_id = student_data.get('grade_id')
                    if not grade_id:
                        continue
                    
                    student_grade = StudentGrade.objects.get(
                        id=grade_id,
                        class_record=class_record
                    )
                    
                    scores = student_data.get('scores', {})
                    
                    # Update WW scores
                    if 'ww' in scores:
                        for i, value in enumerate(scores['ww'], start=1):
                            setattr(student_grade, f'ww_score_{i}', value or 0)
                    
                    # Update PT scores
                    if 'pt' in scores:
                        for i, value in enumerate(scores['pt'], start=1):
                            setattr(student_grade, f'pt_score_{i}', value or 0)
                    
                    # Update QA score
                    if 'qa' in scores:
                        student_grade.qa_score_1 = scores['qa'][0] or 0
                    
                    # Save will auto-calculate grades via model's save() method
                    student_grade.save()
            
            # Log the save action
            log_class_record_action(
                class_record=class_record,
                action='updated',
                user=teacher.user,
                description=f'Updated scores and configuration for {class_record.section.name} - {class_record.subject.subject_name} ({class_record.quarter})'
            )
        
        return JsonResponse({
            'success': True,
            'message': 'Class record saved successfully',
            'timestamp': timezone.now().isoformat()
        })
        
    except Teacher.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Teacher profile not found'
        }, status=404)
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Invalid JSON data'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def log_class_record_action(class_record, action, user, description=''):
    """
    Log actions performed on class records for audit trail.
    This can be expanded to use a dedicated ActivityLog model.
    """
    from admin_functionalities.models import ActivityLog
    
    action_messages = {
        'created': 'Created class record',
        'updated': 'Updated class record',
        'exported_pdf': 'Exported class record to PDF',
        'exported_excel': 'Exported class record to Excel',
        'printed': 'Printed class record',
    }
    
    activity = ActivityLog.objects.create(
        user=user,
        action=action_messages.get(action, action),
        module='Class Record',
        timestamp=timezone.now()
    )
    
    return activity


# ============================================================================
# EXPORT ENDPOINTS - PDF & Excel
# ============================================================================

@login_required
@require_http_methods(["GET"])
def export_class_record_pdf(request, record_id):
    """
    Export class record to PDF format with professional header matching masterlist style.
    Uses ReportLab to generate a formatted PDF document.
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        from reportlab.pdfgen import canvas
        from io import BytesIO
        
        teacher = Teacher.objects.get(user=request.user)
        class_record = get_object_or_404(ClassRecord, id=record_id, teacher=teacher)
        
        # Get student grades
        student_grades = StudentGrade.objects.filter(
            class_record=class_record
        ).select_related('student').order_by('student__last_name', 'student__first_name')
        
        # Create PDF buffer
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer, 
            pagesize=landscape(letter), 
            rightMargin=0.5*inch, 
            leftMargin=0.5*inch,
            topMargin=1.5*inch,  # More space for header
            bottomMargin=0.75*inch
        )
        
        # Container for PDF elements
        elements = []
        styles = getSampleStyleSheet()
        
        # =====================================================================
        # HEADER - Matching Masterlist Style
        # =====================================================================
        
        # Header styles
        header_style = ParagraphStyle(
            'HeaderStyle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.black,
            alignment=TA_CENTER,
            spaceAfter=2,
            leading=12
        )
        
        title_style = ParagraphStyle(
            'TitleStyle',
            parent=styles['Heading1'],
            fontSize=14,
            textColor=colors.black,
            alignment=TA_CENTER,
            spaceAfter=3,
            fontName='Helvetica-Bold',
            leading=16
        )
        
        subtitle_style = ParagraphStyle(
            'SubtitleStyle',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.black,
            alignment=TA_CENTER,
            spaceAfter=2,
            leading=11
        )
        
        section_title_style = ParagraphStyle(
            'SectionTitle',
            parent=styles['Heading2'],
            fontSize=12,
            textColor=colors.black,
            alignment=TA_CENTER,
            spaceAfter=8,
            fontName='Helvetica-Bold',
            leading=14
        )
        
        # Build header
        elements.append(Paragraph("Republic of the Philippines", header_style))
        elements.append(Paragraph("Department of Education", header_style))
        elements.append(Paragraph("REGION IX – ZAMBOANGA PENINSULA", header_style))
        elements.append(Paragraph("Schools Division Office of Zamboanga City", header_style))
        elements.append(Paragraph("Zamboanga National High School West", title_style))
        elements.append(Paragraph("R.T. Lim Boulevard, Zamboanga City, Philippines", subtitle_style))
        elements.append(Paragraph("School ID: 303942", subtitle_style))
        elements.append(Spacer(1, 0.15*inch))
        
        # Section title
        program_display = dict(Section._meta.get_field('program').choices).get(
            class_record.section.program, 
            class_record.section.program
        )
        elements.append(Paragraph(
            f"{program_display}-{class_record.section.name} E-CLASS RECORD",
            section_title_style
        ))
        elements.append(Spacer(1, 0.1*inch))
        
        # Info section
        info_style = ParagraphStyle(
            'InfoStyle',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.black,
            alignment=TA_LEFT,
            spaceAfter=3
        )
        
        elements.append(Paragraph(f"<b>Subject:</b> {class_record.subject.subject_name}", info_style))
        elements.append(Paragraph(f"<b>Section:</b> {class_record.section.name}", info_style))
        elements.append(Paragraph(f"<b>Quarter:</b> {class_record.get_quarter_display()}", info_style))
        elements.append(Paragraph(f"<b>School Year:</b> {class_record.school_year}", info_style))
        elements.append(Spacer(1, 0.15*inch))
        
        # =====================================================================
        # TABLE DATA
        # =====================================================================
        
        # Prepare table data
        data = []
        
        # Header row 1
        header_row_1 = [
            'No.', 'Learners\' Names', 'Sex',
            'WW1', 'WW2', 'WW3', 'WW4', 'WW5', 'WW6', 'WW7', 'WW8', 'WW9', 'WW10',
            'Total', 'PS', 'WS',
            'PT1', 'PT2', 'PT3', 'PT4', 'PT5', 'PT6', 'PT7', 'PT8', 'PT9', 'PT10',
            'Total', 'PS', 'WS',
            'QA', 'PS', 'WS',
            'Initial', 'Quarterly'
        ]
        data.append(header_row_1)
        
        # HPS row
        hps_row = [
            '', 'Highest Possible Score', '',
            class_record.ww_hps_1, class_record.ww_hps_2, class_record.ww_hps_3,
            class_record.ww_hps_4, class_record.ww_hps_5, class_record.ww_hps_6,
            class_record.ww_hps_7, class_record.ww_hps_8, class_record.ww_hps_9,
            class_record.ww_hps_10,
            class_record.get_ww_hps_total(), '100', f'{class_record.written_works_weight}',
            class_record.pt_hps_1, class_record.pt_hps_2, class_record.pt_hps_3,
            class_record.pt_hps_4, class_record.pt_hps_5, class_record.pt_hps_6,
            class_record.pt_hps_7, class_record.pt_hps_8, class_record.pt_hps_9,
            class_record.pt_hps_10,
            class_record.get_pt_hps_total(), '100', f'{class_record.performance_tasks_weight}',
            class_record.qa_hps_1, '100', f'{class_record.quarterly_assessment_weight}',
            '', ''
        ]
        data.append(hps_row)
        
        # Student data rows
        for idx, sg in enumerate(student_grades, start=1):
            row = [
                idx,
                f"{sg.student.last_name}, {sg.student.first_name}",
                sg.student.gender[0] if sg.student.gender else 'M',  # M or F
                
                # WW scores
                sg.ww_score_1 or 0, sg.ww_score_2 or 0, sg.ww_score_3 or 0,
                sg.ww_score_4 or 0, sg.ww_score_5 or 0, sg.ww_score_6 or 0,
                sg.ww_score_7 or 0, sg.ww_score_8 or 0, sg.ww_score_9 or 0,
                sg.ww_score_10 or 0,
                f'{sg.ww_total:.2f}', f'{sg.ww_percentage:.2f}', f'{sg.ww_weighted_score:.2f}',
                
                # PT scores
                sg.pt_score_1 or 0, sg.pt_score_2 or 0, sg.pt_score_3 or 0,
                sg.pt_score_4 or 0, sg.pt_score_5 or 0, sg.pt_score_6 or 0,
                sg.pt_score_7 or 0, sg.pt_score_8 or 0, sg.pt_score_9 or 0,
                sg.pt_score_10 or 0,
                f'{sg.pt_total:.2f}', f'{sg.pt_percentage:.2f}', f'{sg.pt_weighted_score:.2f}',
                
                # QA scores
                sg.qa_score_1 or 0, f'{sg.qa_percentage:.2f}', f'{sg.qa_weighted_score:.2f}',
                
                # Grades
                f'{sg.initial_grade:.2f}', sg.quarterly_grade
            ]
            data.append(row)
        
        # =====================================================================
        # CREATE TABLE WITH STYLING - COMPACT VERSION
        # =====================================================================
        
        # Compact column widths to fit everything on one page
        col_widths = [
            0.22*inch,  # No
            1.4*inch,   # Name (shorter)
            0.25*inch,  # Sex
        ] + [0.27*inch] * 10  # WW1-10 (smaller)
        col_widths += [0.32*inch, 0.32*inch, 0.32*inch]  # WW Total, PS, WS
        col_widths += [0.27*inch] * 10  # PT1-10 (smaller)
        col_widths += [0.32*inch, 0.32*inch, 0.32*inch]  # PT Total, PS, WS
        col_widths += [0.27*inch, 0.32*inch, 0.32*inch]  # QA, PS, WS
        col_widths += [0.35*inch, 0.38*inch]  # Initial, Quarterly
        
        table = Table(data, colWidths=col_widths, repeatRows=2)
        
        # Table styling - more compact
        table_style = TableStyle([
            # Header row
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#D3D3D3')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 6),  # Smaller header font
            ('BOTTOMPADDING', (0, 0), (-1, 0), 4),  # Less padding
            ('TOPPADDING', (0, 0), (-1, 0), 4),
            
            # HPS row
            ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#FFFACD')),
            ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 1), (-1, 1), 6),  # Smaller HPS font
            ('BOTTOMPADDING', (0, 1), (-1, 1), 3),
            ('TOPPADDING', (0, 1), (-1, 1), 3),
            
            # Name column left-aligned
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('LEFTPADDING', (1, 0), (1, -1), 3),
            
            # Grid
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('LINEBELOW', (0, 0), (-1, 0), 1.5, colors.black),
            ('LINEBELOW', (0, 1), (-1, 1), 1, colors.black),
            
            # Data rows - more compact
            ('FONTSIZE', (0, 2), (-1, -1), 5.5),  # Smaller data font
            ('ROWBACKGROUNDS', (0, 2), (-1, -1), [colors.white, colors.HexColor('#F9F9F9')]),
            ('BOTTOMPADDING', (0, 2), (-1, -1), 2),  # Minimal padding
            ('TOPPADDING', (0, 2), (-1, -1), 2),
            ('LEFTPADDING', (0, 2), (-1, -1), 2),
            ('RIGHTPADDING', (0, 2), (-1, -1), 2),
        ])
        
        table.setStyle(table_style)
        elements.append(table)
        
        # =====================================================================
        # FOOTER - Signature
        # =====================================================================
        
        elements.append(Spacer(1, 0.3*inch))
        
        footer_style = ParagraphStyle(
            'FooterStyle',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.black,
            alignment=TA_LEFT,
            spaceAfter=3
        )
        
        elements.append(Paragraph("<b>Prepared by:</b>", footer_style))
        elements.append(Spacer(1, 0.5*inch))
        
        signature_style = ParagraphStyle(
            'SignatureStyle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.black,
            alignment=TA_LEFT,
            spaceAfter=2,
            fontName='Helvetica-Bold'
        )
        
        position_style = ParagraphStyle(
            'PositionStyle',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.black,
            alignment=TA_LEFT
        )
        
        elements.append(Paragraph(f"{teacher.full_name}", signature_style))
        elements.append(Paragraph(f"{teacher.position or 'Teacher II'}", position_style))
        
        # Build PDF
        doc.build(elements)
        
        # Get PDF value
        pdf = buffer.getvalue()
        buffer.close()
        
        # Log export action
        log_class_record_action(
            class_record=class_record,
            action='exported_pdf',
            user=teacher.user,
            description=f'Exported PDF for {class_record.section.name} - {class_record.subject.subject_name}'
        )
        
        # Return PDF response
        filename = f"{class_record.section.program}-{class_record.section.name}_ClassRecord_{class_record.subject.subject_code}_{class_record.quarter}_{class_record.school_year.replace('-', '_')}.pdf"
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response.write(pdf)
        
        return response
        
    except ImportError:
        return JsonResponse({
            'success': False,
            'error': 'ReportLab library not installed. Run: pip install reportlab'
        }, status=500)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@login_required
@require_http_methods(["GET"])
def export_class_record_excel(request, record_id):
    """
    Export class record to Excel format.
    Uses openpyxl to generate a formatted Excel workbook.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        
        teacher = Teacher.objects.get(user=request.user)
        class_record = get_object_or_404(ClassRecord, id=record_id, teacher=teacher)
        
        # Get student grades
        student_grades = StudentGrade.objects.filter(
            class_record=class_record
        ).select_related('student').order_by('student__last_name', 'student__first_name')
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"{class_record.quarter} - {class_record.section.name}"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1e3a8a", end_color="1e3a8a", fill_type="solid")
        hps_fill = PatternFill(start_color="fffacd", end_color="fffacd", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        
        # Title rows
        ws.merge_cells('A1:E1')
        ws['A1'] = "E-CLASS RECORD"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = center_align
        
        ws['A2'] = f"Subject: {class_record.subject.subject_name}"
        ws['A3'] = f"Section: {class_record.section.name}"
        ws['A4'] = f"Quarter: {class_record.get_quarter_display()}"
        ws['A5'] = f"School Year: {class_record.school_year}"
        
        # Headers - Row 7
        row = 7
        col = 1
        
        # Basic columns
        headers_basic = ['No.', 'Name', 'Sex']
        for h in headers_basic:
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_align
            col += 1
        
        # Written Works columns
        for i in range(1, 11):
            cell = ws.cell(row=row, column=col, value=f'WW{i}')
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_align
            col += 1
        
        for h in ['Total', 'PS', 'WS']:
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_align
            col += 1
        
        # Performance Tasks columns
        for i in range(1, 11):
            cell = ws.cell(row=row, column=col, value=f'PT{i}')
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_align
            col += 1
        
        for h in ['Total', 'PS', 'WS']:
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_align
            col += 1
        
        # QA columns
        for h in ['QA', 'PS', 'WS']:
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_align
            col += 1
        
        # Grade columns
        for h in ['Initial Grade', 'Quarterly Grade']:
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_align
            col += 1
        
        # HPS Row - Row 8
        row = 8
        col = 1
        
        ws.cell(row=row, column=col, value='').fill = hps_fill
        col += 1
        ws.cell(row=row, column=col, value='Highest Possible Score').fill = hps_fill
        col += 1
        ws.cell(row=row, column=col, value='').fill = hps_fill
        col += 1
        
        # WW HPS
        hps_ww = [class_record.ww_hps_1, class_record.ww_hps_2, class_record.ww_hps_3,
                  class_record.ww_hps_4, class_record.ww_hps_5, class_record.ww_hps_6,
                  class_record.ww_hps_7, class_record.ww_hps_8, class_record.ww_hps_9,
                  class_record.ww_hps_10]
        for hps in hps_ww:
            cell = ws.cell(row=row, column=col, value=hps)
            cell.fill = hps_fill
            cell.alignment = center_align
            col += 1
        
        ws.cell(row=row, column=col, value=class_record.get_ww_hps_total()).fill = hps_fill
        col += 1
        ws.cell(row=row, column=col, value=100).fill = hps_fill
        col += 1
        ws.cell(row=row, column=col, value=class_record.written_works_weight).fill = hps_fill
        col += 1
        
        # PT HPS
        hps_pt = [class_record.pt_hps_1, class_record.pt_hps_2, class_record.pt_hps_3,
                  class_record.pt_hps_4, class_record.pt_hps_5, class_record.pt_hps_6,
                  class_record.pt_hps_7, class_record.pt_hps_8, class_record.pt_hps_9,
                  class_record.pt_hps_10]
        for hps in hps_pt:
            cell = ws.cell(row=row, column=col, value=hps)
            cell.fill = hps_fill
            cell.alignment = center_align
            col += 1
        
        ws.cell(row=row, column=col, value=class_record.get_pt_hps_total()).fill = hps_fill
        col += 1
        ws.cell(row=row, column=col, value=100).fill = hps_fill
        col += 1
        ws.cell(row=row, column=col, value=class_record.performance_tasks_weight).fill = hps_fill
        col += 1
        
        # QA HPS
        ws.cell(row=row, column=col, value=class_record.qa_hps_1).fill = hps_fill
        col += 1
        ws.cell(row=row, column=col, value=100).fill = hps_fill
        col += 1
        ws.cell(row=row, column=col, value=class_record.quarterly_assessment_weight).fill = hps_fill
        col += 1
        
        # Empty cells for grades
        ws.cell(row=row, column=col, value='').fill = hps_fill
        col += 1
        ws.cell(row=row, column=col, value='').fill = hps_fill
        
        # Student data rows
        row = 9
        for idx, sg in enumerate(student_grades, start=1):
            col = 1
            
            ws.cell(row=row, column=col, value=idx).alignment = center_align
            col += 1
            ws.cell(row=row, column=col, value=f"{sg.student.last_name}, {sg.student.first_name}")
            col += 1
            ws.cell(row=row, column=col, value=sg.student.gender).alignment = center_align
            col += 1
            
            # WW scores
            for score in sg.get_ww_scores_list():
                ws.cell(row=row, column=col, value=score).alignment = center_align
                col += 1
            ws.cell(row=row, column=col, value=round(sg.ww_total, 2)).alignment = center_align
            col += 1
            ws.cell(row=row, column=col, value=round(sg.ww_percentage, 2)).alignment = center_align
            col += 1
            ws.cell(row=row, column=col, value=round(sg.ww_weighted_score, 2)).alignment = center_align
            col += 1
            
            # PT scores
            for score in sg.get_pt_scores_list():
                ws.cell(row=row, column=col, value=score).alignment = center_align
                col += 1
            ws.cell(row=row, column=col, value=round(sg.pt_total, 2)).alignment = center_align
            col += 1
            ws.cell(row=row, column=col, value=round(sg.pt_percentage, 2)).alignment = center_align
            col += 1
            ws.cell(row=row, column=col, value=round(sg.pt_weighted_score, 2)).alignment = center_align
            col += 1
            
            # QA scores
            ws.cell(row=row, column=col, value=sg.qa_score_1).alignment = center_align
            col += 1
            ws.cell(row=row, column=col, value=round(sg.qa_percentage, 2)).alignment = center_align
            col += 1
            ws.cell(row=row, column=col, value=round(sg.qa_weighted_score, 2)).alignment = center_align
            col += 1
            
            # Grades
            ws.cell(row=row, column=col, value=round(sg.initial_grade, 2)).alignment = center_align
            col += 1
            ws.cell(row=row, column=col, value=sg.quarterly_grade).alignment = center_align
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 5
        
        # Save to BytesIO
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Log export action
        log_class_record_action(
            class_record=class_record,
            action='exported_excel',
            user=teacher.user,
            description=f'Exported Excel for {class_record.section.name} - {class_record.subject.subject_name}'
        )
        
        # Return Excel response
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="ClassRecord_{class_record.section.name}_{class_record.subject.subject_code}_{class_record.quarter}.xlsx"'
        
        return response
        
    except ImportError:
        return JsonResponse({
            'success': False,
            'error': 'openpyxl library not installed. Run: pip install openpyxl'
        }, status=500)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


# ============================================================================
# HISTORY & LOGS
# ============================================================================

@login_required
@require_http_methods(["GET"])
def get_class_record_history(request):
    """
    Get activity history for class records.
    Shows recent saves, exports, and modifications.
    """
    try:
        from admin_functionalities.models import ActivityLog
        
        teacher = Teacher.objects.get(user=request.user)
        
        # Get recent logs for this teacher's class record activities
        logs = ActivityLog.objects.filter(
            user=teacher.user,
            module='Class Record'
        ).order_by('-timestamp')[:20]
        
        history_data = [{
            'id': log.id,
            'action': log.action,
            'date': log.date,
            'time': log.time,
            'timestamp': log.timestamp.isoformat()
        } for log in logs]
        
        return JsonResponse({
            'success': True,
            'history': history_data
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
        
# =========================FOR THE WARNINGS==================================================
@login_required
@require_http_methods(["POST"])
def check_early_warnings(request):
    """
    Check for students at risk and create/update early warning alerts.
    Called after grades are calculated.
    FIXED: Check initial grade instead of quarterly grade for warnings.
    """
    try:
        teacher = Teacher.objects.get(user=request.user)
        data = json.loads(request.body)
        
        class_record_id = data.get('class_record_id')
        quarter = data.get('quarter')
        
        class_record = get_object_or_404(ClassRecord, id=class_record_id, teacher=teacher)
        
        warnings_created = []
        warnings_resolved = []
        
        # Get all student grades for this class record
        student_grades = StudentGrade.objects.filter(class_record=class_record)
        
        print(f"\n{'='*80}")
        print(f"CHECKING EARLY WARNINGS - {class_record.section.name} {quarter}")
        print(f"{'='*80}")
        
        for sg in student_grades:
            # ====================================================================
            # CRITICAL FIX: Check INITIAL GRADE, not quarterly grade
            # Initial grade of 60 = Quarterly grade of 75 (passing)
            # So we need to check if initial grade < 60
            # ====================================================================
            
            initial_grade = sg.initial_grade  # Use this instead!
            quarterly_grade = sg.quarterly_grade  # This is just for display
            
            print(f"\nStudent: {sg.student.last_name}, {sg.student.first_name}")
            print(f"  Initial Grade: {initial_grade:.2f}")
            print(f"  Quarterly Grade (Transmuted): {quarterly_grade}")
            
            # Skip if no scores entered yet
            if initial_grade == 0:
                print(f"  → Skipping (no scores entered)")
                continue
            
            # Check if QA is complete
            qa_complete = sg.qa_score_1 > 0
            
            # ================================================================
            # WARNING THRESHOLDS (based on INITIAL grade, not transmuted)
            # ================================================================
            # Initial Grade >= 60 → Transmutes to 75+ → PASSING
            # Initial Grade 56-59 → Transmutes to 74  → AT RISK
            # Initial Grade < 56  → Transmutes to <74 → AT RISK
            # Initial Grade < 40  → Transmutes to <70 → HIGH RISK
            # ================================================================
            
            warning_type = None
            
            if qa_complete:
                # Quarter is complete - final assessment
                if initial_grade < 60:  # Will transmute to <75
                    warning_type = 'failed'
                    print(f"  → FAILED (QA complete, initial grade {initial_grade:.2f} < 60)")
                else:
                    print(f"  → PASSING (QA complete, initial grade {initial_grade:.2f} >= 60)")
            else:
                # Quarter in progress - projection
                if initial_grade < 50:  # Very low - high risk
                    warning_type = 'at_risk'
                    print(f"  → AT HIGH RISK (initial grade {initial_grade:.2f} < 50, QA pending)")
                elif initial_grade < 60:  # Below passing threshold
                    warning_type = 'at_risk'
                    print(f"  → AT RISK (initial grade {initial_grade:.2f} < 60, QA pending)")
                else:
                    print(f"  → ON TRACK (initial grade {initial_grade:.2f} >= 60)")
            
            if warning_type:
                # Calculate required performance
                required = calculate_required_performance(sg, class_record)
                
                print(f"  → Creating/Updating warning: {warning_type}")
                print(f"     Points needed: {required['points_needed']}")
                
                # Check if warning already exists
                existing_warning = EarlyWarningAlert.objects.filter(
                    student_grade=sg,
                    class_record=class_record,
                    status='active'
                ).first()
                
                if existing_warning:
                    # Update existing warning
                    existing_warning.warning_type = warning_type
                    existing_warning.current_grade = quarterly_grade  # Display grade
                    existing_warning.initial_grade = initial_grade     # Actual grade
                    existing_warning.required_performance = required
                    existing_warning.save()
                    print(f"  → Updated existing warning #{existing_warning.id}")
                else:
                    # Create new warning
                    warning = EarlyWarningAlert.objects.create(
                        student_grade=sg,
                        student=sg.student,
                        class_record=class_record,
                        warning_type=warning_type,
                        current_grade=quarterly_grade,  # Display grade
                        initial_grade=initial_grade,     # Actual grade  
                        required_performance=required
                    )
                    warnings_created.append({
                        'student_id': sg.student.id,
                        'student_name': f"{sg.student.last_name}, {sg.student.first_name}",
                        'warning_type': warning_type,
                        'initial_grade': initial_grade,
                        'quarterly_grade': quarterly_grade
                    })
                    print(f"  → Created new warning #{warning.id}")
            else:
                # Student is on track - resolve any active warnings
                resolved = EarlyWarningAlert.objects.filter(
                    student_grade=sg,
                    status='active'
                ).update(status='resolved', resolved_at=timezone.now())
                
                if resolved > 0:
                    print(f"  → Resolved {resolved} warning(s)")
                    warnings_resolved.append(sg.student.id)
        
        print(f"\n{'='*80}")
        print(f"WARNINGS SUMMARY:")
        print(f"  Created: {len(warnings_created)}")
        print(f"  Resolved: {len(warnings_resolved)}")
        print(f"{'='*80}\n")
        
        return JsonResponse({
            'success': True,
            'warnings_created': warnings_created,
            'warnings_resolved': len(warnings_resolved),
            'total_active': EarlyWarningAlert.objects.filter(
                class_record=class_record,
                status='active'
            ).count()
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


def calculate_required_performance(student_grade, class_record):
    """
    Calculate what scores student needs on remaining tasks to reach passing.
    FIXED: Use proper percentage calculations based on INITIAL grade.
    """
    # Get weights
    ww_weight = class_record.written_works_weight / 100
    pt_weight = class_record.performance_tasks_weight / 100
    qa_weight = class_record.quarterly_assessment_weight / 100
    
    # Get HPS totals
    ww_hps_total = class_record.get_ww_hps_total()
    pt_hps_total = class_record.get_pt_hps_total()
    qa_hps_total = class_record.get_qa_hps_total()
    
    # Current status
    current_ww_ws = student_grade.ww_weighted_score
    current_pt_ws = student_grade.pt_weighted_score
    current_qa_ws = student_grade.qa_weighted_score
    current_initial_grade = student_grade.initial_grade
    
    # Check if QA is done
    qa_done = student_grade.qa_score_1 > 0
    
    # Target: Initial Grade = 60 → Transmutes to 75
    target_initial_grade = 60.0
    points_needed = max(0, target_initial_grade - current_initial_grade)
    
    result = {
        'current_grade': student_grade.quarterly_grade,
        'initial_grade': current_initial_grade,
        'target_grade': 75,
        'target_initial_grade': target_initial_grade,
        'points_needed': round(points_needed, 2),
        'qa_completed': qa_done,
        'warning_type': 'failed' if qa_done else 'at_risk'
    }
    
    # ============================================================
    # CASE 1: QA NOT DONE (Quarter In Progress)
    # ============================================================
    if not qa_done:
        # Get scores
        ww_scores = [
            student_grade.ww_score_1, student_grade.ww_score_2, student_grade.ww_score_3,
            student_grade.ww_score_4, student_grade.ww_score_5, student_grade.ww_score_6,
            student_grade.ww_score_7, student_grade.ww_score_8, student_grade.ww_score_9,
            student_grade.ww_score_10
        ]
        pt_scores = [
            student_grade.pt_score_1, student_grade.pt_score_2, student_grade.pt_score_3,
            student_grade.pt_score_4, student_grade.pt_score_5, student_grade.pt_score_6,
            student_grade.pt_score_7, student_grade.pt_score_8, student_grade.pt_score_9,
            student_grade.pt_score_10
        ]
        
        ww_completed = [s for s in ww_scores if s > 0]
        pt_completed = [s for s in pt_scores if s > 0]
        
        ww_current_avg = sum(ww_completed) / len(ww_completed) if ww_completed else 0
        pt_current_avg = sum(pt_completed) / len(pt_completed) if pt_completed else 0
        
        # Calculate needed percentages
        ww_ps_current = student_grade.ww_percentage
        ww_ps_needed = ((target_initial_grade * ww_weight) / ww_weight) if ww_weight > 0 else 0
        ww_improvement = max(0, ww_ps_needed - ww_ps_current)
        
        pt_ps_current = student_grade.pt_percentage
        pt_ps_needed = ((target_initial_grade * pt_weight) / pt_weight) if pt_weight > 0 else 0
        pt_improvement = max(0, pt_ps_needed - pt_ps_current)
        
        qa_ps_needed = ((target_initial_grade * qa_weight) / qa_weight) if qa_weight > 0 else 0
        qa_score_needed = (qa_ps_needed / 100) * qa_hps_total
        
        result['advice'] = {
            'ww': {
                'current_avg': round(ww_current_avg, 1),
                'current_percentage': round(ww_ps_current, 1),
                'needed_percentage': round(min(100, ww_ps_needed), 1),
                'improvement_needed': round(ww_improvement, 1),
                'message': (
                    f"Maintain at least {min(100, ww_ps_needed):.0f}% in Written Works" 
                    if ww_improvement > 0 
                    else "Current Written Works performance is adequate"
                )
            },
            'pt': {
                'current_avg': round(pt_current_avg, 1),
                'current_percentage': round(pt_ps_current, 1),
                'needed_percentage': round(min(100, pt_ps_needed), 1),
                'improvement_needed': round(pt_improvement, 1),
                'message': (
                    f"Maintain at least {min(100, pt_ps_needed):.0f}% in Performance Tasks"
                    if pt_improvement > 0
                    else "Current Performance Tasks performance is adequate"
                )
            },
            'qa': {
                'needed_percentage': round(min(100, qa_ps_needed), 1),
                'needed_score': round(min(qa_hps_total, max(0, qa_score_needed)), 1),
                'message': f"Score at least {min(qa_hps_total, max(0, qa_score_needed)):.0f}/{qa_hps_total} on the Quarterly Assessment"
            },
            'overall_message': (
                f"Student needs {points_needed:.1f} more points to reach passing grade."
                if points_needed > 0
                else "Student is on track to pass."
            )
        }
    
    # ============================================================
    # CASE 2: QA DONE (Quarter Complete)
    # ============================================================
    else:
        result['advice'] = {
            'message': f"Quarter is complete. Student scored {student_grade.quarterly_grade} (initial grade: {current_initial_grade:.1f}, below passing threshold of 60).",
            'intervention_required': True,
            'components': {
                'ww': {
                    'percentage': round(student_grade.ww_percentage, 1),
                    'weighted': round(student_grade.ww_weighted_score, 2)
                },
                'pt': {
                    'percentage': round(student_grade.pt_percentage, 1),
                    'weighted': round(student_grade.pt_weighted_score, 2)
                },
                'qa': {
                    'percentage': round(student_grade.qa_percentage, 1),
                    'weighted': round(student_grade.qa_weighted_score, 2)
                }
            }
        }
    
    return result


def calculate_required_performance(student_grade, class_record):
    """
    Calculate what scores student needs on remaining tasks to reach 75.
    Returns dict with required scores for each component.
    """
    # Get weights
    ww_weight = class_record.written_works_weight / 100
    pt_weight = class_record.performance_tasks_weight / 100
    qa_weight = class_record.quarterly_assessment_weight / 100
    
    # Get HPS totals
    ww_hps_total = class_record.get_ww_hps_total()
    pt_hps_total = class_record.get_pt_hps_total()
    qa_hps_total = class_record.get_qa_hps_total()
    
    # Count completed tasks (non-zero scores)
    ww_scores = student_grade.get_ww_scores_list()
    pt_scores = student_grade.get_pt_scores_list()
    
    ww_completed = sum(1 for score in ww_scores if score > 0)
    pt_completed = sum(1 for score in pt_scores if score > 0)
    qa_completed = 1 if student_grade.qa_score_1 > 0 else 0
    
    ww_remaining = 10 - ww_completed
    pt_remaining = 10 - pt_completed
    qa_remaining = 1 - qa_completed
    
    # Calculate current weighted scores
    current_ww_ws = student_grade.ww_weighted_score
    current_pt_ws = student_grade.pt_weighted_score
    current_qa_ws = student_grade.qa_weighted_score
    
    # Target initial grade (75 transmutes from ~60)
    target_initial_grade = 60.0
    
    # Calculate points needed
    points_needed = target_initial_grade - (current_ww_ws + current_pt_ws + current_qa_ws)
    
    # Simple approach: distribute evenly across remaining tasks
    result = {
        'ww_remaining_count': ww_remaining,
        'pt_remaining_count': pt_remaining,
        'qa_remaining_count': qa_remaining,
        'points_needed': round(points_needed, 2),
    }
    
    # Calculate minimum needed on each remaining task
    if ww_remaining > 0:
        # How many points needed from WW component
        ww_points_needed = (target_initial_grade * ww_weight) - current_ww_ws
        # What percentage needed on remaining WW
        ww_percentage_needed = (ww_points_needed / ww_weight) / ww_remaining if ww_remaining > 0 else 0
        # Convert to score out of 10 (assuming each WW is worth 10)
        result['ww_min_per_task'] = min(10, max(0, round(ww_percentage_needed / 10, 1)))
    
    if pt_remaining > 0:
        pt_points_needed = (target_initial_grade * pt_weight) - current_pt_ws
        pt_percentage_needed = (pt_points_needed / pt_weight) / pt_remaining if pt_remaining > 0 else 0
        result['pt_min_per_task'] = min(10, max(0, round(pt_percentage_needed / 10, 1)))
    
    if qa_remaining > 0:
        qa_points_needed = (target_initial_grade * qa_weight) - current_qa_ws
        qa_percentage_needed = (qa_points_needed / qa_weight)
        result['qa_min_score'] = min(qa_hps_total, max(0, round((qa_percentage_needed / 100) * qa_hps_total, 1)))
    
    return result


@login_required
@require_http_methods(["GET"])
def get_early_warnings(request):
    """
    Get all active early warnings for a class record.
    """
    try:
        teacher = Teacher.objects.get(user=request.user)
        class_record_id = request.GET.get('class_record_id')
        
        class_record = get_object_or_404(ClassRecord, id=class_record_id, teacher=teacher)
        
        warnings = EarlyWarningAlert.objects.filter(
            class_record=class_record,
            status='active'
        ).select_related('student', 'student_grade')
        
        warnings_data = [{
            'id': w.id,
            'student_id': w.student.id,
            'student_name': f"{w.student.last_name}, {w.student.first_name}",
            'warning_type': w.warning_type,
            'current_grade': w.current_grade,
            'required_performance': w.required_performance,
            'intervention_enabled': w.warning_type == 'failed',
            'flagged_at': w.flagged_at.isoformat()
        } for w in warnings]
        
        return JsonResponse({
            'success': True,
            'warnings': warnings_data
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@require_http_methods(["POST"])
def dismiss_warning(request):
    """
    Dismiss an early warning alert.
    """
    try:
        data = json.loads(request.body)
        warning_id = data.get('warning_id')
        notes = data.get('notes', '')
        
        warning = get_object_or_404(EarlyWarningAlert, id=warning_id)
        warning.dismiss(notes)
        
        return JsonResponse({
            'success': True,
            'message': 'Warning dismissed'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)